#!/usr/bin/env python
# coding: utf-8

# In[1]:


import cx_Oracle
from openpyxl import load_workbook

def create_oracle_connection(db_params):
    try:
        dsn_tns = cx_Oracle.makedsn(db_params['host'], db_params['port'], service_name=db_params['service_name'])
        connection = cx_Oracle.connect(user=db_params['username'], password=db_params['password'], dsn=dsn_tns)
        print("Oracle Database connection established successfully.")
        return connection
    except cx_Oracle.DatabaseError as e:
        print(f"Error connecting to Oracle Database: {e}")
        return None

def table_exists(connection, table_name):
    try:
        cursor = connection.cursor()
        cursor.execute(f"SELECT COUNT(*) FROM user_tables WHERE table_name = '{table_name.upper()}'")
        table_exists = cursor.fetchone()[0] == 1
        cursor.close()
        return table_exists
    except cx_Oracle.DatabaseError as e:
        error, = e.args
        print(f"Error checking if table exists: {error.message}")
        return False

def create_table(connection, table_name, columns):
    try:
        cursor = connection.cursor()
        columns_def = ', '.join([f'"{col.upper()}" VARCHAR2(4000)' for col in columns])  # Default to VARCHAR2(4000) for simplicity
        create_table_sql = f"CREATE TABLE {table_name.upper()} ({columns_def})"
        cursor.execute(create_table_sql)
        cursor.close()
        connection.commit()
        print(f"Table '{table_name}' created successfully.")
    except cx_Oracle.DatabaseError as e:
        error, = e.args
        print(f"Error creating table: {error.message}")

def insert_data(connection, table_name, columns, data):
    try:
        cursor = connection.cursor()
        placeholders = ', '.join([':' + str(i+1) for i in range(len(columns))])
        insert_columns = ', '.join([f'"{col.upper()}"' for col in columns])
        insert_sql = f'INSERT INTO {table_name.upper()} ({insert_columns}) VALUES ({placeholders})'
        
        for i, row in enumerate(data):
            try:
                cursor.execute(insert_sql, row)
            except cx_Oracle.DatabaseError as e:
                error, = e.args
                print(f"Error inserting row {i}: {error.message} - Row data: {row}")
        
        cursor.close()
        connection.commit()
        print("Data inserted successfully.")
    except cx_Oracle.DatabaseError as e:
        error, = e.args
        print(f"Error inserting data: {error.message}")

def update_or_insert_data(connection, table_name, columns, data, primary_key):
    try:
        cursor = connection.cursor()

        # Prepare SQL statements
        placeholders = ', '.join([':' + str(i+1) for i in range(len(columns))])
        update_columns = ', '.join([f'"{col.upper()}" = :{i+1}' for i, col in enumerate(columns)])
        insert_columns = ', '.join([f'"{col.upper()}"' for col in columns])
        insert_sql = f'INSERT INTO {table_name.upper()} ({insert_columns}) VALUES ({placeholders})'
        update_sql = f'UPDATE {table_name.upper()} SET {update_columns} WHERE "{primary_key.upper()}" = :{len(columns)+1}'

        for i, row in enumerate(data):
            try:
                # Check if row exists for update
                cursor.execute(f'SELECT COUNT(*) FROM {table_name.upper()} WHERE "{primary_key.upper()}" = :1', (row[columns.index(primary_key)],))
                row_count = cursor.fetchone()[0]
                
                if row_count > 0:
                    cursor.execute(update_sql, row + [row[columns.index(primary_key)]])
                else:
                    cursor.execute(insert_sql, row)
            except cx_Oracle.DatabaseError as e:
                error, = e.args
                print(f"Error updating or inserting row {i}: {error.message} - Row data: {row}")
        
        cursor.close()
        connection.commit()
        print("Data updated/inserted successfully.")
    except cx_Oracle.DatabaseError as e:
        error, = e.args
        print(f"Error updating/inserting data: {error.message}")

def load_excel_data_into_oracle(db_params, table_name, excel_file_path, primary_key):
    try:
        wb = load_workbook(excel_file_path)
        ws = wb.active
        print("Excel file read successfully.")

        # Extract column names
        columns = [cell.value.upper() for cell in ws[1]]

        # Extract data rows
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append([str(cell) if cell is not None else '' for cell in row])
        
        print("Columns in the Excel file:", columns)
        print("First few rows of data:", data[:5])

        connection = create_oracle_connection(db_params)
                 
        if connection:
            if table_exists(connection, table_name):
                print(f"Table '{table_name}' already exists. Updating/inserting data...")
                update_or_insert_data(connection, table_name, columns, data, primary_key)
            else:
                print(f"Table '{table_name}' does not exist. Creating and inserting data.")
                create_table(connection, table_name, columns)
                insert_data(connection, table_name, columns, data)
            
            connection.close()
            print("Oracle Database connection closed.")
    
    except FileNotFoundError as e:
        print(f"Error: {e}")
    except KeyError as e:
        print(f"KeyError: {e}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

# Database parameters for Oracle
db_params = {
    'username': 'SYSTEM',
    'password': '1234',
    'host': 'DESKTOP-CCM4VHF',
    'port': '1521',
    'service_name': 'XE'
}

# Define the table name and the path to the Excel file
table_name = 'weather_details'
excel_file_path = r'C:\Users\wissen\Downloads\weather_data.xlsx'
primary_key = 'CITY'  # Assuming 'CITY' is the primary key

# Remove any unwanted Unicode characters
excel_file_path = excel_file_path.replace('\u202a', '').replace('\u202b', '')

# Call the function to create the table and ingest data
load_excel_data_into_oracle(db_params, table_name, excel_file_path, primary_key)


# In[ ]:




